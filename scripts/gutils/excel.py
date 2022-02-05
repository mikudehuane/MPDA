# -*- coding: utf-8 -*-
# @Time    : 2021/7/30 上午11:02
# @Author  : islander
# @File    : excel.py
# @Software: PyCharm
from copy import deepcopy
from itertools import chain

from openpyxl.utils import get_column_letter, coordinate_to_tuple


def get_dimension(sheet):
    """获取 sheet 的范围，用 int 表示

    Args:
        sheet: 待处理的工作表

    Returns: 范围，均右包含，以 int 表示
        行范围最小值，行范围最大值, 列范围最小值，列范围最大值
    """
    dimensions = sheet.dimensions

    left_top, right_bottom = dimensions.split(':')
    top, left = coordinate_to_tuple(left_top)
    bottom, right = coordinate_to_tuple(right_bottom)

    return left, right, top, bottom


def tuple_to_coordinate(row, col):  # 将行列表示的内容转为 excel 的 cell 名

    col = get_column_letter(col)
    row = str(row)
    return '{}{}'.format(col, row)


def append_col(sheet, col):  # 在表后添加一列内容，返回与 cell 坐标的列表，如 ['K1', 'K2', 'K3']
    if sheet.max_column == 1:
        # 判断第一列是否为空
        if sheet.max_row == 1:  # 只有一个元素
            if sheet['A1'].value is None:
                max_column = 0
            else:
                max_column = 1
        else:
            max_column = 1
    else:
        max_column = sheet.max_column
    col_to_insert = max_column + 1

    cell_coords = []
    if len(col) > 0:
        for row, cell_val in enumerate(col, 1):
            coord = tuple_to_coordinate(row, col_to_insert)
            cell_coords.append(coord)
            sheet[coord] = cell_val
    else:  # 创建空列
        sheet.cell(row=1, column=col_to_insert)
        sheet.cell(row=2, column=col_to_insert)  # 创建两行，如果只创建一行，会在只有一列时被认为是空表

    return cell_coords


def append_row(sheet, row, *, convert_types=True, row_idx=None, col_start=None):
    """添加一行到工作表

    Args:
        col_start (int): 追加行的第一个元素所在列，如果不指定，则从首个非 None 列开始添加
        row_idx: 默认在表后追加一行，如果指定 row_idx，则在指定行的数据后追加（忽略 None 数据）
        row: 待添加的行
        sheet: 待添加的工作表
        convert_types: 是否将字符串类型转为合适的类别
    """
    if not convert_types:
        values = row
    else:  # 转换类别
        values = []
        for entry in row:
            if isinstance(entry, str):  # 尝试转为浮点数或者整数
                if entry.isdigit():
                    entry = int(entry)
                else:
                    try:
                        entry = float(entry)
                    except ValueError:
                        pass
            values.append(entry)

    if row_idx is None:
        sheet.append(values)
    else:
        if col_start is None:
            exist_row = sheet[row_idx]
            exist_row = get_valid_cells(exist_row)
            last_col, _ = coordinate_to_tuple(exist_row[-1].coordinate)
            current_column = last_col + 1  # 当前列为已经存在的列后的一个列
        else:
            current_column = col_start  # 从指定的列开始填

        for cell_value in values:
            sheet.cell(row=row_idx, column=current_column).value = cell_value
            current_column += 1


def get_valid_cells(cells):
    """从一个 cell 数组提取有效的 cell 范围（去掉右边为 None 的 cell）

    Args:
        cells: cell 的数组

    Returns:
        删除对应元素后的 cell 数组
    """
    cells = [x for x in cells]
    # 从尾部开始弹出 None cell
    pop_count = 0
    for cell in cells[::-1]:
        if cell.value is None:
            pop_count += 1
        else:
            break
    [cells.pop() for _ in range(pop_count)]
    return cells
