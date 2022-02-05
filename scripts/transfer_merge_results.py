# -*- coding: utf-8 -*-
# @Time    : 2021/7/28 下午3:15
# @Author  : islander
# @File    : transfer_merge_results.py
# @Software: PyCharm


import logging

import argparse
import pickle
import re
import time
from collections import OrderedDict
from copy import deepcopy
from itertools import chain

import openpyxl.cell.cell
from openpyxl.utils import get_column_letter, coordinate_to_tuple
from tensorflow.python.platform import gfile

import log.logging_config
from gutils import parse_fp, CachedActions, excel, get_ordered_dict
import os.path as osp
from openpyxl import Workbook
import pandas as pd


_logger = logging.getLogger('transfer')
_custom_logger = log.CustomLogger(logger=_logger)

_CELL_PAT = '{}!{}'
# 所有可能出现的表键的顺序
_SHEET_KEY_ORDER = ['local', 'init', 'selected', 'both', 'match']


def _get_col_range(sheet, col_name):
    """获取一列内容的范围，按照 excel cell 格式返回起点和终点，如 'A1'， 'A92'，会从底部去掉为 None 的 cell
    """
    left, right, top, bottom = excel.get_dimension(sheet)
    heading = [cell.value for cell in sheet[1]]
    col_name2idx = {col_name: idx for idx, col_name in enumerate(heading, 1)}  # 列名与 excel 中列号的对应
    col_idx = col_name2idx[col_name]
    column = sheet[get_column_letter(col_idx)]  # 目标的列cells
    column = excel.get_valid_cells(column)  # 提取出有效的元素
    bottom, _ = coordinate_to_tuple(column[-1].coordinate)
    start = excel.tuple_to_coordinate(2, col_idx)
    end = excel.tuple_to_coordinate(bottom, col_idx)
    return start, end


def merge_into_array(fds):
    """将 fds 指定的路径中的内容融合到一些列表（二维）中

    Args:
        fds: 需要被处理的目录的列表

    Returns:
        tuple[Dict[str, list], Dict[str, list]]:
            tuple(trainset_result, testset_result)，每个 result 都是一个字典，键是运行的 key，值是融合的结果，第一行是表头
    """

    def _append_for_arr(_arr_dict, _key, _row, _other_msg):  # 将一行内容添加到数组 _arr 中，并在第一行添加表头
        """将一行数据添加到数组中，并在第一行添加表头

        Args:
            _arr_dict: 待被填充的数组的字典
            _key: 被填充的数组对应的键
            _row: DataFrame 的行对象，会被填入 _arr
            _other_msg (OrderedDict): 希望附加的信息，会被填入 _arr 的前几列
        """
        if _key not in _arr_dict:
            _arr_dict[_key] = []
        _arr = _arr_dict[_key]

        if not _arr:  # 没有写表头
            _heading = list(_other_msg.keys()) + list(_row.keys())
            _arr.append(_heading)
        _values = list(_other_msg.values()) + list(_row.values)
        _arr.append(_values)

    def append_for_file(fp, arr_dict, _other_msg):
        """将 fp 路径文件的内容整理到 arr_dict 对应的数组中

        Args:
            _other_msg (OrderedDict): 希望附加的信息，会被填入 _arr 的前几列
            arr_dict: 被填充的数组，因一个文件对应多个数组，以字典形式传入
            fp: 待解析的文件，解析内容被填入数组
        """
        if gfile.Exists(fp):
            # noinspection PyTypeChecker
            fp_data = pd.read_csv(gfile.GFile(fp), dtype=str)

            _selected_row = None  # 外部数据训练出的模型的行
            for idx, row in fp_data.iterrows():
                if row.step == 'local_data_trained':
                    append_action.add_action(_append_for_arr, args=(arr_dict, 'local', row, _other_msg))
                elif row.step == 'init_model':
                    append_action.add_action(_append_for_arr, args=(arr_dict, 'init', row, _other_msg))
                    _selected_row = row  # 如果什么数据都没选择，那初始模型就是选择的模型
                elif row.step == 'local&external_data_trained':
                    append_action.add_action(_append_for_arr, args=(arr_dict, 'both', row, _other_msg))
                elif row.step == 'vanilla_match':
                    append_action.add_action(_append_for_arr, args=(arr_dict, 'match', row, _other_msg))
                else:  # 迭代中的值
                    if row.is_current_user_selected.lower() == 'true':  # 被选择了
                        _selected_row = row
            append_action.add_action(_append_for_arr, args=(arr_dict, 'selected', _selected_row, _other_msg))

    append_action = CachedActions()

    # 存放结果的缓存
    trainset_fs = OrderedDict()
    testset_fs = OrderedDict()

    for fd_idx, fd in enumerate(fds):
        user_id = osp.split(fd)[-1]
        _logger.info('processing directory {} ({}/{} processed)'.format(fd, fd_idx, len(fds)))

        train_fp = osp.join(fd, 'trainset.csv')  # 训练集结果的绝对路径
        test_fp = osp.join(fd, 'testset.csv')  # 测试集评估结果的绝对路径

        other_msg = OrderedDict()
        other_msg['user_id'] = user_id
        append_for_file(train_fp, trainset_fs, other_msg)
        append_for_file(test_fp, testset_fs, other_msg)

        # TODO(islander): 当增加实验时，这里要相应增加
        if append_action.num_actions < 10:
            _custom_logger.log_text(
                'WARNING: file {} or {} is not complete, the run is not valid'.format(train_fp, test_fp))
            append_action.clear_actions()
        else:
            append_action.apply_actions()
            append_action.clear_actions()

        # 添加的顺序与需要的顺序不一定一致，这里更新顺序
        trainset_fs = get_ordered_dict(trainset_fs, [key for key in _SHEET_KEY_ORDER if key in trainset_fs])
        testset_fs = get_ordered_dict(testset_fs, [key for key in _SHEET_KEY_ORDER if key in testset_fs])

    return trainset_fs, testset_fs


def merge_arrs_from_workers(arrays, *, sorted_by_uid=True):
    """将多个 worker 提取出的 ndarray 拼接成一个

    Args:
        arrays: 各个 worker 产出的数组的列表，每一项都是 train_arrs, test_arrs 元组，均为字典，字典值为第一行是表头的数据
        sorted_by_uid: 是否按 int(user_id) 排序拼接后的数组，默认每一行第一项为 user_id

    Returns:
        与 arrays[0] 格式相同的 train_arrs, test_arrs 元组
    """
    trainset_full_arrs = dict()
    testset_full_arrs = dict()
    for trainset_arrs, testset_arrs in arrays:

        def _fill_into(_src, _dest):
            for _key, _val in _src.items():
                if _key not in _dest:
                    _dest[_key] = []
                    _dest[_key].extend(_val)  # 包括表头
                else:
                    _dest[_key].extend(_val[1:])  # 不包括表头

        _fill_into(trainset_arrs, trainset_full_arrs)
        _fill_into(testset_arrs, testset_full_arrs)

    if sorted_by_uid:  # 按用户排序

        def _sort(_arrs):
            for _val in _arrs.values():
                _val[1:] = sorted(_val[1:], key=lambda x: int(x[0]))

        _sort(trainset_full_arrs)
        _sort(testset_full_arrs)

    return trainset_full_arrs, testset_full_arrs


def insert_arr_to_worksheet(arr, worksheet, *, convert_types=True):
    """将 python 数组写入工作表

    Args:
        convert_types: 是否将字符串类型转为合适的类别
        worksheet: 待添加的工作表
        arr: 待添加的数组
    """
    for row in arr:
        excel.append_row(worksheet, row, convert_types=convert_types)


# 将 numpy 对象转为 workbook 对象
def convert_arrays_to_workbook(arrays, *, add_formula=True):
    """将 arrays 对象转为 WorkBook 对象，并添加各种公式

    Args:
        add_formula: 是否添加公式信息，如计算 GAUC 的公式
        arrays: (train_arrays, test_arrays) 两个字典的元组

    Returns:
        添加了所有信息的 WorkBook 对象
    """

    def add_sheet(sheet_dict, key, prefix):  # 向字典添加一个工作表并填写表头，sheet_name 为键，加了前缀为 excel 中 sheet 名
        _sheet = result_f.create_sheet('{}_{}'.format(prefix, key))
        sheet_dict[key] = _sheet

    def get_sheet_dict(prefix, _sheet_keys):  # 向字典添加所有几个工作表
        sheet_dict = OrderedDict()
        # 添加所有几个工作表
        [add_sheet(sheet_dict, key, prefix) for key in _sheet_keys]
        return sheet_dict

    def _append_arrays_to_sheets(_arrays, _sheets):  # 将 _arrays 按键匹配添加到 _sheets
        for _key, _array in _arrays.items():
            _sheet = _sheets[_key]
            insert_arr_to_worksheet(arr=_array, worksheet=_sheet)

    def get_col_ref_and_add(_col_name, _col_index, _sheet_name):
        """在 merged_sheet 中添加一列，引用 _sheet 的数据

        Args:
            _sheet_name: 被引用工作表的名称
            _col_name: 引用列的名称
            _col_index: 被引用列的 index 'A' 'B' 等
        """
        _sheet_dict_name, _sheet_key = _sheet_name.split('_')
        if _sheet_dict_name == 'trainset':
            _sheet = trainset_fs[_sheet_key]
        elif _sheet_dict_name == 'testset':
            _sheet = testset_fs[_sheet_key]
        else:
            raise ValueError('_sheet_dict_name={}'.format(_sheet_dict_name))

        _ref_col = [_col_name]
        _reffed_col = _sheet[_col_index]
        for cell in _reffed_col[1:]:
            ref_str = '={}!{}'.format(_sheet_name, cell.coordinate)
            _ref_col.append(ref_str)
        excel.append_col(merged_sheet, _ref_col)

    def _get_col_range_str(_sheet, _col_name='auc'):  # 获取 auc 对应的范围，附加表标题
        _start, _end = _get_col_range(_sheet, _col_name)
        return '{}:{}'.format(_start, _end)

    def _append_cmp_count_for_train_or_test(_sheet_dict, _set_name):  # 填充数据，避免训练集和测试集代码重复
        _ranges = {
            key: _get_col_range_str(merged_sheet, '{}_{}'.format(_set_name, key))
            for key in _sheet_dict.keys()
        }

        def _append(_row_name, _formula_pat):
            _row_to_append = [_row_name]
            for _key in _sheet_dict.keys():
                if _key != 'init':
                    _row_to_append.append(_formula_pat.format(_ranges[_key], _ranges['init']))
            _append_row(merged_sheet, _row_to_append)

        _append('增长', '=SUMPRODUCT(--({}>{}))')
        _append('不变', '=SUMPRODUCT(--({}={}))')
        _append('下降', '=SUMPRODUCT(--({}<{}))')

    trainset_arrs, testset_arrs = arrays

    # 创建工作表对象
    result_f = Workbook()
    result_f.remove_sheet(result_f.active)  # 删除初始的表
    trainset_fs = get_sheet_dict('trainset', _sheet_keys=trainset_arrs.keys())
    testset_fs = get_sheet_dict('testset', _sheet_keys=testset_arrs.keys())

    _append_arrays_to_sheets(trainset_arrs, trainset_fs)
    _append_arrays_to_sheets(testset_arrs, testset_fs)

    # 写附加信息
    if add_formula:
        merged_sheet = result_f.create_sheet('merged_result', index=0)
        # 引用其他表格的 user_id, num_selected_users, AUC
        num_samples_col = 'K'
        get_col_ref_and_add('user_id', _col_index='A', _sheet_name='trainset_local')
        get_col_ref_and_add('num_selected_users', _col_index='D', _sheet_name='trainset_selected')
        get_col_ref_and_add('num_train_samples', _col_index=num_samples_col, _sheet_name='trainset_local')
        get_col_ref_and_add('num_test_samples', _col_index=num_samples_col, _sheet_name='testset_local')
        # 添加所有工作表的内容
        for sheet in chain(trainset_fs.values(), testset_fs.values()):
            sheet_name = sheet.title
            get_col_ref_and_add(sheet_name, _col_index='F', _sheet_name=sheet_name)
        excel.append_col(merged_sheet, [''] * merged_sheet.max_row)  # 空一列, max_row 为当前 row 的数量

        # 推理当前列数
        formula_start_column = merged_sheet.max_column + 1  # 公式的第一列，放的是行名，如 'local'

        current_row = [1]  # 为了能在内嵌函数中自增

        def _append_row(*_args, **_kwargs):  # 在 current_row 追加内容，并自增该列，用于一致化与 append 的调用
            excel.append_row(*_args, **_kwargs,
                             row_idx=current_row[0], convert_types=False, col_start=formula_start_column)
            current_row[0] += 1

        # 填充 GAUC 或者不加权的 AUC 均值，_formula_temp 为公式字符串的格式化模板
        def _append_for_gauc_or_avgauc(_formula_temp):
            # 计算 init 和 local 对应行的索引
            _init_row_idx = current_row[0] + 1
            _local_row_idx = current_row[0]
            _current_start_row = current_row[0]

            # 通过 zip 实现"转置"，一行行写
            for row_idx, ((train_sheet_key, train_sheet), (test_sheet_key, test_sheet)) in enumerate(
                    zip(trainset_fs.items(), testset_fs.items()), _current_start_row):

                def _get_gauc_string(_sheet_name, _num_samples_col_name):  # 获取计算 gauc 的公式

                    _auc_start, _auc_end = _get_col_range(merged_sheet, _sheet_name)
                    _num_samples_start, _num_samples_end = _get_col_range(merged_sheet, _num_samples_col_name)

                    _formula = _formula_temp.format(auc_start=_auc_start, auc_end=_auc_end,
                                                    num_samples_start=_num_samples_start, num_samples_end=_num_samples_end)
                    return _formula

                train_gauc = _get_gauc_string(train_sheet.title, 'num_train_samples')
                test_gauc = _get_gauc_string(test_sheet.title, 'num_test_samples')

                train_col = get_column_letter(formula_start_column + 1)
                test_col = get_column_letter(formula_start_column + 2)
                _append_row(
                    merged_sheet,
                    [train_sheet_key, train_gauc, test_gauc,
                     '={}{}/{}{}'.format(train_col, row_idx, train_col, _init_row_idx),
                     '={}{}/{}{}'.format(train_col, row_idx, train_col, _local_row_idx),  # 训练集比值
                     '={}{}/{}{}'.format(test_col, row_idx, test_col, _init_row_idx),
                     '={}{}/{}{}'.format(test_col, row_idx, test_col, _local_row_idx)])  # 测试集比值

        # 写 GAUC 公式
        _append_row(merged_sheet,
                    [None, '训练集 GAUC', '测试集 GAUC',
                     '与初始模型相比的比例（训练集）', '与本地训练相比的比例（训练集）',
                     '与初始模型相比的比例（测试集）', '与本地训练相比的比例（测试集）'],
                    )  # 表头
        _append_for_gauc_or_avgauc(
            '=SUMPRODUCT(({auc_start}:{auc_end}>=0)*{auc_start}:{auc_end},{num_samples_start}:{num_samples_end})'
            '/'
            'SUMIF({auc_start}:{auc_end},">=0",{num_samples_start}:{num_samples_end})'
        )

        current_row[0] += 1  # 空行

        # 写平均 GAUC 公式
        _append_row(merged_sheet,
                    [None, '训练集平均 AUC', '测试集平均 GAUC',
                     '与初始模型相比的比例（训练集）', '与本地训练相比的比例（训练集）',
                     '与初始模型相比的比例（测试集）', '与本地训练相比的比例（测试集）'],
                    )  # 表头
        _append_for_gauc_or_avgauc(
            '=SUMIF({auc_start}:{auc_end},">=0")'
            '/'
            'COUNTIF({auc_start}:{auc_end},">=0")'
        )

        # 统计增长降低不变的用户数
        current_row[0] += 1  # 空行
        _append_row(merged_sheet, ['训练集相比初始模型 xxx 的用户数',
                                   *[key for key in trainset_arrs.keys() if key != 'init']])
        _append_cmp_count_for_train_or_test(trainset_fs, 'trainset')
        _append_row(merged_sheet, ['测试集集相比初始模型 xxx 的用户数',
                                   *[key for key in testset_arrs.keys() if key != 'init']])
        _append_cmp_count_for_train_or_test(testset_fs, 'testset')

        # 统计选择的用户数
        current_row[0] += 1  # 空行
        num_selected_users_range = _get_col_range_str(merged_sheet, 'num_selected_users')
        _append_row(merged_sheet, ['平均被选择的用户数', '=AVERAGE({})'.format(num_selected_users_range)])

    return result_f


def main():
    parser = argparse.ArgumentParser(description='将 oss 上下载的数据转为 xlsx 表格文件')
    parser.add_argument('-fp', '--result_fp', type=parse_fp, required=True,
                        help='oss 下载文件的路径，或给定一个目录，转化其下所有 pkl 后缀的文件')
    args, unparsed_args = parser.parse_known_args()

    if gfile.IsDirectory(args.result_fp):
        fns = gfile.ListDirectory(args.result_fp)  # 目录下所有文件
        fns = [fn for fn in fns if fn.rsplit('.', 1)[-1] == 'pkl']  # 过滤留下 pkl 文件
        fps = [osp.join(args.result_fp, fn) for fn in fns]  # 转为绝对路径
    else:
        fps = [args.result_fp]

    for fp in fps:
        # noinspection PyTypeChecker
        full_arr = pickle.load(gfile.GFile(fp, 'rb'))

        # 转为 workbook
        workbook = convert_arrays_to_workbook(full_arr)

        # 生成输出文件的名称
        dest_fn = osp.split(fp)[-1]
        prefix, suffix = dest_fn.rsplit('.', 1)
        dest_fn = '.'.join([prefix, 'xlsx'])

        dest_fp = osp.join(args.result_fp, dest_fn)
        workbook.save(dest_fp)

        _logger.info('convert {} to {}'.format(fp, dest_fp))


if __name__ == '__main__':
    main()
